"""Tests for billing.exports — Excel workbook generation."""

import pytest

from billing.exports import (
    build_database_workbook,
    build_enrollments_sheet,
    build_payments_sheet,
    build_students_sheet,
)

pytestmark = pytest.mark.django_db


class TestBuildStudentsSheet:
    def test_header_row(self, db):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        build_students_sheet(ws)
        headers = [cell.value for cell in ws[1]]
        assert "Nombre" in headers
        assert "Apellidos" in headers
        assert "Grupo" in headers
        assert "Tutor - DNI" in headers

    def test_includes_student_data(self, student_with_parent):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        build_students_sheet(ws)
        # Row 1 = header, Row 2 = student
        assert ws.max_row == 2
        assert ws.cell(row=2, column=2).value == "Lucas"
        assert ws.cell(row=2, column=3).value == "López García"


class TestBuildEnrollmentsSheet:
    def test_includes_enrollment_data(self, active_enrollment):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        build_enrollments_sheet(ws)
        assert ws.max_row == 2
        assert ws.cell(row=2, column=5).value == "2025-2026"

    def test_empty_when_no_enrollments(self, db):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        build_enrollments_sheet(ws)
        assert ws.max_row == 1  # Only header


class TestBuildPaymentsSheet:
    def test_includes_payment_data(self, pending_payment):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        build_payments_sheet(ws)
        assert ws.max_row == 2
        assert ws.cell(row=2, column=8).value == "54.00"


class TestBuildDatabaseWorkbook:
    def test_three_sheets(self, student_with_parent, active_enrollment, pending_payment):
        wb = build_database_workbook()
        assert len(wb.sheetnames) == 3
        assert wb.sheetnames == ["Estudiantes", "Matrículas", "Pagos"]

    def test_empty_database(self, db):
        wb = build_database_workbook()
        assert len(wb.sheetnames) == 3
        # All sheets should have header only
        for ws in wb.worksheets:
            assert ws.max_row == 1
