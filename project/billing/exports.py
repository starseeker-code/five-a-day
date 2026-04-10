"""
exports.py — Reusable data export helpers.

Each function returns an openpyxl Workbook (or a single Worksheet) so callers
can decide how to deliver it (HTTP response, save to file, attach to email, …).
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from billing.models import Enrollment, Payment
from students.models import Student


_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="EC4899", end_color="EC4899", fill_type="solid")
_CENTER = Alignment(horizontal="center", vertical="center")


def _style_header(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)


def _d(d):
    return d.strftime("%d/%m/%Y") if d else ""


def build_students_sheet(ws):
    _style_header(ws, [
        "ID", "Nombre", "Apellidos", "Fecha Nacimiento", "Colegio", "Alergias",
        "RGPD Firmado", "Grupo", "Activo", "Fecha Baja", "Motivo Baja",
        "Tutor - Nombre", "Tutor - Apellidos", "Tutor - DNI",
        "Tutor - Teléfono", "Tutor - Email", "Tutor - IBAN", "Fecha Alta",
    ])
    qs = Student.objects.select_related("group").prefetch_related("parents").order_by("last_name", "first_name")
    for s in qs:
        parents = list(s.parents.all())
        ws.append([
            s.id, s.first_name, s.last_name, _d(s.birth_date), s.school, s.allergies,
            "Sí" if s.gdpr_signed else "No", s.group.group_name if s.group else "",
            "Sí" if s.active else "No", _d(s.withdrawal_date), s.withdrawal_reason,
            " / ".join(p.first_name for p in parents), " / ".join(p.last_name for p in parents),
            " / ".join(p.dni for p in parents), " / ".join(p.phone for p in parents),
            " / ".join(p.email for p in parents), " / ".join(p.iban for p in parents),
            _d(s.created_at),
        ])
    _auto_width(ws)


def build_enrollments_sheet(ws):
    _style_header(ws, [
        "ID", "Estudiante - Nombre", "Estudiante - Apellidos",
        "Tipo Matrícula", "Año Académico", "Tipo Horario",
        "Inicio Período", "Fin Período", "Fecha Matrícula",
        "Importe Base", "Descuento %", "Importe Final",
        "Estado", "URL Documento", "Notas", "Fecha Creación",
    ])
    qs = Enrollment.objects.select_related("student", "enrollment_type").order_by("-enrollment_date")
    for e in qs:
        ws.append([
            e.id, e.student.first_name, e.student.last_name,
            e.enrollment_type.display_name, e.academic_year, e.get_schedule_type_display(),
            _d(e.enrollment_period_start), _d(e.enrollment_period_end), _d(e.enrollment_date),
            float(e.enrollment_amount), float(e.discount_percentage), float(e.final_amount),
            e.get_status_display(), e.document_url, e.notes, _d(e.created_at),
        ])
    _auto_width(ws)


def build_payments_sheet(ws):
    _style_header(ws, [
        "ID", "Estudiante - Nombre", "Estudiante - Apellidos",
        "Tutor - Nombre", "Tutor - Apellidos", "Tutor - DNI",
        "Concepto", "Importe", "Moneda", "Tipo Pago", "Método Pago",
        "Estado", "Fecha Vencimiento", "Fecha Pago",
        "Referencia", "Observaciones", "Fecha Creación",
    ])
    qs = Payment.objects.select_related("student", "parent").order_by("-due_date")
    for p in qs:
        ws.append([
            p.id, p.student.first_name, p.student.last_name,
            p.parent.first_name, p.parent.last_name, p.parent.dni,
            p.concept, float(p.amount), p.currency,
            p.get_payment_type_display(), p.get_payment_method_display(),
            p.get_payment_status_display(), _d(p.due_date), _d(p.payment_date),
            p.reference_number, p.observations, _d(p.created_at),
        ])
    _auto_width(ws)


def build_database_workbook():
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Estudiantes"
    build_students_sheet(ws1)
    ws2 = wb.create_sheet("Matrículas")
    build_enrollments_sheet(ws2)
    ws3 = wb.create_sheet("Pagos")
    build_payments_sheet(ws3)
    return wb
